##################################
#Author: Eddie Sanchez 
#Date: May 10, 2019
#Assignment: Evulatation of openpyxl 
#Final Project Python Script 2
#Description of Script this script takes the values from the pervious script (Batch_Processing_Files_Script1.py) and  add the total arces
#This Script is for EPIC, The Environmental Protection Information Center, They asked me to calcuate the total arces of for each of the re-occuring Timber Identification Numbers (THP_NUM)
#The Out Come of this script is having one Identification number (HD_NUM) [located in Column R in the excel files] with one total arces printed to the same excel file under 'Total Arces" Under Column "V"
######################################

#documentation used for openpyxl: https://www.geeksforgeeks.org/python-arithmetic-operations-in-excel-file-using-openpyxl/
#documentation used for pands: https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html

##########################################
#Uses Anconda to process files 
#to run code first use Anaconda C:-> Program Files-> Anaconda3 -> python.exe
#########################################

#Libraries Used
import openpyxl
import pandas as pd 
import openpyxl as px

################################################################
#DelNorte2016 Excel File Manipulation 
##################################################################
#array methods used to maninpate excel spread sheet  
#An array is a list used to store temporarl used to define and excute in script 

######################################################
#Set up File Path 
Outputpath="C:\\Users\\ees407\\Desktop\\Python Final Project\\Outputs\\"

data_file=pd.read_excel(Outputpath+'DelNorte2016.xlsx') #used to read the file 

############################################
#set up Outputpath for Jim to make it easier to process

Outputpath="C:\\Users\\ees407\\Desktop\\Python Final Project\\Outputs\\"

##########################################
#define arces

acres=data_file['ACRES'] #Arces is defined to be used throughout the script 

######################################
# define the Spilt 

num= data_file['THP_NUM'] #spilt is used to know where the last digits of the identification number changes

#import the THP_NUM into the first list 
nums= list(data_file['THP_NUM'])

#########################################
#first array
my_list= []
count_list=[]
for x in nums: 
    try: 
        index=my_list.index(x)
    except:
        count_list.append(0)
        my_list.append(x)

#print(my_list)

index=0
##################################
#second array 
while (index<len(nums)): 
    a1=nums[index]
    
    #print(a1)
    b1= acres[index]
    #print(b1)
    
    index2=my_list.index(a1)
    count_list[index2]+=b1
   # print(count_list)
    index+=1
#########################
#loop that goes through both array 
for i in count_list: 
    try: 
        index=count_list.index(i)
    except:
        my_list.append(0)
        count_list.append(i)

       
#print(count_list)

index=0

#line of code that prints in columns 
LineCount=0
wb=openpyxl.load_workbook(Outputpath+"DelNorte2016.xlsx")
TheFile=(Outputpath+"DelNorte2016.xlsx","w")
sheet= wb.active
sheet=wb.get_sheet_by_name('DelNorte2016')
sheet["V1"] = ("Total Arces")
sheet["V2"] = (format(count_list))
print(count_list)
count=0
for areacount in count_list: 
    thecell=("V"+format(count+2))
    
    try:
        sheet[thecell]=count_list[int(count)]
    except:
        pass
    
    count+=1
    
#save work using xl
    
wb.save(Outputpath+"DelNorte2016.xlsx")


###############################

###############################

##DelNorte2017 Processing for the second shapefile  
#################################
#Outputpath="C:\\Users\\ees407\\Desktop\\Python Final Project\\Outputs\\"

data_file=pd.read_excel(Outputpath+'DelNorte2017.xlsx')

acres=data_file['ACRES']

######################################
# First Array

num= data_file['THP_NUM']

nums= list(data_file['THP_NUM'])
#create new array 
my_list= []
count_list=[]
for x in nums: 
    try: 
        index=my_list.index(x)
    except:
        count_list.append(0)
        my_list.append(x)

#print(my_list)

index=0
##################
#second array 
while (index<len(nums)): 
    a1=nums[index]
    
    #print(a1)
    b1= acres[index]
    #print(b1)
    
    index2=my_list.index(a1)
    count_list[index2]+=b1
   # print(count_list)
    
    
    index+=1
#########################
#loop that goes through both array 
for i in count_list: 
    try: 
        index=count_list.index(i)
    except:
        my_list.append(0)
        count_list.append(i)

       
print(count_list)

index=0

#line of code that prints in columns 
LineCount=0
wb=openpyxl.load_workbook(Outputpath+"DelNorte2017.xlsx")
TheFile=(Outputpath+"DelNorte2017.xlsx","w")
sheet= wb.active
sheet=wb.get_sheet_by_name('DelNorte2017')
sheet["V1"] = ("Total Arces")
sheet["V2"] = (format(count_list))
print(count_list)
count=0
for areacount in count_list: 
    thecell=("V"+format(count+2))
    
    try:
        sheet[thecell]=count_list[int(count)]
    except:
        pass
    
    count+=1
    
wb.save(Outputpath+"DelNorte2017.xlsx")


###################################
#####################################
##DelNorte2018

######################################

#Outputpath="C:\\Users\\ees407\\Desktop\\Python Final Project\\Outputs\\"

data_file=pd.read_excel(Outputpath+'DelNorte2018.xlsx')

acres=data_file['ACRES']

######################################
# Jim help here 

num= data_file['THP_NUM']

nums= list(data_file['THP_NUM'])
#create new array 
my_list= []
count_list=[]
for x in nums: 
    try: 
        index=my_list.index(x)
    except:
        count_list.append(0)
        my_list.append(x)

#print(my_list)

index=0
##################
#second array 
while (index<len(nums)): 
    a1=nums[index]
    
    #print(a1)
    b1= acres[index]
    #print(b1)
    
    index2=my_list.index(a1)
    count_list[index2]+=b1
   # print(count_list)
    
    
    index+=1
#########################
#loop that goes through both array 
for i in count_list: 
    try: 
        index=count_list.index(i)
    except:
        my_list.append(0)
        count_list.append(i)

       
print(count_list)

index=0


#line of code that prints in columns 
LineCount=0
wb=openpyxl.load_workbook(Outputpath+"DelNorte2018.xlsx")
TheFile=(Outputpath+"DelNorte2018.xlsx","w")
sheet= wb.active
sheet=wb.get_sheet_by_name('DelNorte2018')
sheet["V1"] = ("Total Arces")
sheet["V2"] = (format(count_list))
print(count_list)
count=0
for areacount in count_list: 
    thecell=("V"+format(count+2))
    
    try:
        sheet[thecell]=count_list[int(count)]
    except:
        pass
    
    count+=1
    

wb.save(Outputpath+"DelNorte2018.xlsx")

####################################
####################################
#Humboldt2016
####################################

#Outputpath="C:\\Users\\ees407.HSU-AD.000\\Desktop\\Python Final Project\\Outputs\\"

data_file=pd.read_excel(Outputpath+'Humboldt2016.xlsx')

#print(num[0:369])# number of cells in the cloumns
acres=data_file['ACRES']
#print(acres[0:3])

######################################
num= data_file['THP_NUM']

nums= list(data_file['THP_NUM'])
#create new array 
my_list= []
count_list=[]
for x in nums: 
    try: 
        index=my_list.index(x)
    except:
        count_list.append(0)
        my_list.append(x)

#print(my_list)

index=0
##################
#second array 
while (index<len(nums)): 
    a1=nums[index]
    
    #print(a1)
    b1= acres[index]
    #print(b1)
    
    index2=my_list.index(a1)
    count_list[index2]+=b1
   # print(count_list)
    
    
    index+=1
#########################
#loop that goes through both array 
for i in count_list: 
    try: 
        index=count_list.index(i)
    except:
        my_list.append(0)
        count_list.append(i)

       
print(count_list)

index=0

#line of code that prints in columns 
LineCount=0
wb=openpyxl.load_workbook(Outputpath+"Humboldt2016.xlsx")
TheFile=(Outputpath+"Humboldt2016.xlsx","w")
sheet= wb.active
sheet=wb.get_sheet_by_name('Humboldt2016')
sheet["V1"] = ("Total Arces")
sheet["V2"] = (format(count_list))
print(count_list)
count=0
for areacount in count_list: 
    thecell=("V"+format(count+2))
    
    try:
        sheet[thecell]=count_list[int(count)]
    except:
        pass
    
    count+=1
    
wb.save(Outputpath+"Humboldt2016.xlsx")

#################################

#Humboldt2017
###################################
#Outputpath="C:\\Users\\ees407\\Desktop\\Python Final Project\\Outputs\\"

data_file=pd.read_excel(Outputpath+'Humboldt2017.xlsx')

acres=data_file['ACRES']


######################################

num= data_file['THP_NUM']

nums= list(data_file['THP_NUM'])
#create new array 
my_list= []
count_list=[]
for x in nums: 
    try: 
        index=my_list.index(x)
    except:
        count_list.append(0)
        my_list.append(x)

#print(my_list)

index=0
##################
#second array 
while (index<len(nums)): 
    a1=nums[index]
    
    #print(a1)
    b1= acres[index]
    #print(b1)
    
    index2=my_list.index(a1)
    count_list[index2]+=b1
   # print(count_list)
    
    
    index+=1
#########################
    
   
#loop that goes through both array 
for i in count_list: 
    try: 
        index=count_list.index(i)
    except:
        my_list.append(0)
        count_list.append(i)

       
print(count_list)

index=0

####################################
#line of code that prints in columns 
LineCount=0
wb=openpyxl.load_workbook(Outputpath+"Humboldt2017.xlsx")
TheFile=(Outputpath+"Humboldt2017.xlsx","w")
sheet= wb.active
sheet=wb.get_sheet_by_name('Humboldt2017')
sheet["V1"] = ("Total Arces")
sheet["V2"] = (format(count_list))
print(count_list)
count=0
for areacount in count_list: 
    thecell=("V"+format(count+2))
    
    try:
        sheet[thecell]=count_list[int(count)]
    except:
        pass
    
    count+=1

wb.save(Outputpath+"Humboldt2017.xlsx")

####################################

#Humboldt2018
#########################################

#Outputpath="C:\\Users\\ees407\\Desktop\\Python Final Project\\Outputs\\"

data_file=pd.read_excel(Outputpath+'Humboldt2018.xlsx')

#print(num[0:369])# number of cells in the cloumns
acres=data_file['ACRES']
#print(acres[0:3])


######################################
num= data_file['THP_NUM']

nums= list(data_file['THP_NUM'])
#create new array 
my_list= []
count_list=[]
for x in nums: 
    try: 
        index=my_list.index(x)
    except:
        count_list.append(0)
        my_list.append(x)

#print(my_list)

index=0
##################
#second array 
while (index<len(nums)): 
    a1=nums[index]
    
    #print(a1)
    b1= acres[index]
    #print(b1)
    
    index2=my_list.index(a1)
    count_list[index2]+=b1
   # print(count_list)
    
    
    index+=1
#########################
#loop that goes through both array 
for i in count_list: 
    try: 
        index=count_list.index(i)
    except:
        my_list.append(0)
        count_list.append(i)

       
print(count_list)

index=0

######################################
#line of code that prints in columns 
LineCount=0
wb=openpyxl.load_workbook(Outputpath+"Humboldt2018.xlsx")
TheFile=(Outputpath+"Humboldt2018.xlsx","w")
sheet= wb.active
sheet=wb.get_sheet_by_name('Humboldt2018')
sheet["V1"] = ("Total Arces")
sheet["V2"] = (format(count_list))
print(count_list)
count=0
for areacount in count_list: 
    thecell=("V"+format(count+2))
    
    try:
        sheet[thecell]=count_list[int(count)]
    except:
        pass
    
    count+=1


wb.save(Outputpath+"Humboldt2018.xlsx")

########################################
#the results should be in the Total Arces Column for each individual Excel File 